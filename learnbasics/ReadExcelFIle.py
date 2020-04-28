import openpyxl as xl

file_path = "../TestFiles - Excel/TestExcel.xlsx"

#open as workbook
workbook = xl.load_workbook(file_path)

#get the sheets
sheet = workbook.active #if we have only one workbook.

#or, We have grab sheets by its' name.
#sheet workbook.get_sheet_by_name("sheet1")

#now we need to find out the number of rows and columns in the sheet.
number_of_rows = sheet.max_row
number_of_cols = sheet.max_column

print("number of rows : {}.".format(number_of_rows))
print("Number of columns : {}.".format(number_of_cols))

for r in range(1, number_of_rows+1):
    for c in range(1, number_of_cols+1 ):
        print(sheet.cell(row=r, column=c).value,end =" | ")
    print()