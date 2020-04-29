import openpyxl as xl


class ExcelUtils:
    def __init__(self, filepath =""):
        if filepath != "":
            self.filepath = filepath
            self.__wb = xl.load_workbook(filepath)
            print("Workbook is loaded.")
            print("creating sheet object..")
            self.__sheet = self.__wb.active
        else:
            print("File path is empty. Please provide the valid path.")

    def get_number_of_rows(self):
        return self.__sheet.max_row

    def get_number_of_columns(self):
        return self.__sheet.max_column

    def read_xl_data(self,row, column):
        try:
            return self.__sheet.cell(row=row, column=column).value
        except:
            print("Please check, if row and column values are passed properly.")
            return

    def update_xl_data(self, row, column, data):
        self.__sheet.cell(row=row, column=column).value = data

    def save_xl_data(self):
        self.__wb.save(self.filepath)